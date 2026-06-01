import sys
import os
import json
import datetime
import openpyxl

EXCEL_PATH = r"c:\Users\Ari\Desktop\Proyectos\medistats\RHB2024 DEFINITIVA ESPINOZAa.xlsx"

COL_MAPPING = {
    "fecha": 1,
    "profesional": 2,
    "fichaUsuario": 3,
    "rutUsuario": 4,
    "diagnostico": 5,
    "codigo": 6,
    "clasificacionPatologia": 7,
    "sexo": 8,
    "edad": 9,
    "mai": 10,
    "pensionado": 11,
    "noBeneficiario": 12,
    "atencionCerrada": 13,
    "atencionAbierta": 14,
    "atencionUrgencia": 15,
    "ingreso": 16,
    "pti": 17,
    "consultaInicial": 18,
    "consultaIntermedia": 19,
    "sesiones": 20,
    
    # Procedures
    "rehabilitacionAuditiva": 56,          # REHABILITACIÓN AUDITIVA INDIVIDUAL
    "audiometriaAdultos": 85,
    "audiometriaNinos": 86,
    "audiometriaCampoLibre": 87,
    "calibracionAudifonos": 88,
    "funcionTubaria": 89,
    "emisionesOtoacusticas": 90,
    "rehabilitacionAuditiva3": 91,         # REHABILITACIÓN AUDITIVA INDIVIDUAL3
    "potencialesEvocadosAuditivos": 92,
    "otoscopia": 93,
    "rehabilitacionVestibular": 94,
    "rehabilitacionLecturaLabiofacial": 95,
    "impendanciometria": 96,
    "pruebaDeAudifonos": 97,
    "testDeGlicerol": 98,
    "pruebaCalorica": 99,
    "examenFuncionalVIIIPar": 100,        # Examen funcional de VIII Par
    "estudioFuncionalEquilibrio": 101,    # Estudio funcional de equilibrio en plataforma
    "timpanometriaBandaAncha": 102,
    "habituacionVestibular": 103,
    "maniobraReposicionParticulas": 104,   # Maniobra de reposción de partículas
    "maniobrasReposicionVestibular": 105,  # Maniobras de reposición vestibular
    "alta": 192
}

def to_bool(val):
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return val == 1
    if isinstance(val, str):
        return val.strip() in ("1", "true", "True", "x", "X")
    return bool(val)

def read_records(n=50):
    if not os.path.exists(EXCEL_PATH):
        return {"error": "Excel file not found at " + EXCEL_PATH}
    
    try:
        # Load data_only=True to get evaluated values of formulas
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        sheet = wb["ESTADISTICA"]
        
        records = []
        count = 0
        
        # Scan from bottom to top
        for r in range(sheet.max_row, 6, -1):
            date_val = sheet.cell(row=r, column=1).value
            prof_val = sheet.cell(row=r, column=2).value
            
            # We want rows that have a date and are Ariel Espinoza
            if date_val is not None and prof_val is not None and ('ARIEL' in str(prof_val).upper() or 'ESPINOZA' in str(prof_val).upper()):
                if isinstance(date_val, (datetime.datetime, datetime.date)):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val)
                
                rec = {
                    "id": str(r),
                    "row": r,
                    "fecha": date_str,
                    "profesional": str(prof_val),
                    "fichaUsuario": str(sheet.cell(row=r, column=3).value or ""),
                    "rutUsuario": str(sheet.cell(row=r, column=4).value or ""),
                    "diagnostico": str(sheet.cell(row=r, column=5).value or ""),
                    "codigo": str(sheet.cell(row=r, column=6).value or ""),
                    "clasificacionPatologia": str(sheet.cell(row=r, column=7).value or ""),
                    "sexo": str(sheet.cell(row=r, column=8).value or ""),
                    "edad": str(sheet.cell(row=r, column=9).value or ""),
                    "mai": to_bool(sheet.cell(row=r, column=10).value),
                    "pensionado": to_bool(sheet.cell(row=r, column=11).value),
                    "noBeneficiario": to_bool(sheet.cell(row=r, column=12).value),
                    "atencionCerrada": to_bool(sheet.cell(row=r, column=13).value),
                    "atencionAbierta": to_bool(sheet.cell(row=r, column=14).value),
                    "atencionUrgencia": to_bool(sheet.cell(row=r, column=15).value),
                    "ingreso": to_bool(sheet.cell(row=r, column=16).value),
                    "pti": to_bool(sheet.cell(row=r, column=17).value),
                    "consultaInicial": to_bool(sheet.cell(row=r, column=18).value),
                    "consultaIntermedia": to_bool(sheet.cell(row=r, column=19).value),
                    "sesiones": int(sheet.cell(row=r, column=20).value or 0),
                }
                
                # Load procedures
                for key, col in COL_MAPPING.items():
                    if col > 20: # Procedure columns
                        rec[key] = to_bool(sheet.cell(row=r, column=col).value)
                
                records.append(rec)
                count += 1
                if count >= n:
                    break
        
        return {"records": records}
    except Exception as e:
        return {"error": str(e)}

def search_patient(query):
    if not os.path.exists(EXCEL_PATH):
        return {"error": "Excel file not found"}
    
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        sheet = wb["USUARIOS"]
        
        query = str(query).strip().lower()
        if not query:
            return {"patients": []}
            
        results = []
        for r in range(2, sheet.max_row + 1):
            ficha = sheet.cell(row=r, column=1).value
            rut = sheet.cell(row=r, column=2).value
            
            ficha_str = str(ficha).strip() if ficha is not None else ""
            rut_str = str(rut).strip() if rut is not None else ""
            
            if query == ficha_str.lower() or query == rut_str.lower().replace(".", "").replace("-", "") or query == rut_str.lower():
                dob = sheet.cell(row=r, column=6).value
                dob_str = ""
                if isinstance(dob, (datetime.datetime, datetime.date)):
                    dob_str = dob.strftime('%Y-%m-%d')
                
                results.append({
                    "fichaUsuario": ficha_str,
                    "rutUsuario": rut_str,
                    "diagnostico": str(sheet.cell(row=r, column=3).value or ""),
                    "sexo": str(sheet.cell(row=r, column=4).value or ""),
                    "edad": str(sheet.cell(row=r, column=5).value or ""),
                    "fechaNacimiento": dob_str
                })
                # Break early since ficha/rut are unique
                break
                
        return {"patients": results}
    except Exception as e:
        return {"error": str(e)}

def write_record(record_data):
    if not os.path.exists(EXCEL_PATH):
        return {"error": "Excel file not found at " + EXCEL_PATH}
        
    try:
        # Load data_only=False to preserve existing formulas!
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
        sheet = wb["ESTADISTICA"]
        
        # Find first row where FECHA (col 1) is empty AND PROFESIONAL (col 2) is "ARIEL ESPINOZA"
        target_row = None
        for r in range(7, sheet.max_row + 1):
            date_val = sheet.cell(row=r, column=1).value
            prof_val = sheet.cell(row=r, column=2).value
            if date_val is None and prof_val is not None and "ARIEL" in str(prof_val).upper():
                target_row = r
                break
        
        is_new_row = False
        if target_row is None:
            # Append new row
            target_row = sheet.max_row + 1
            is_new_row = True
            
            # Copy formatting and formulas from row 2990 or a nearby row if possible
            # Let's find a representative template row
            template_row = 2990 if sheet.max_row >= 2990 else 7
            
            # Copy cells
            for col in range(1, 219):
                cell_template = sheet.cell(row=template_row, column=col)
                cell_new = sheet.cell(row=target_row, column=col)
                if cell_template.value is not None:
                    val_str = str(cell_template.value)
                    if val_str.startswith("="):
                        # Copy formula
                        cell_new.value = val_str
                    elif col == 2:
                        cell_new.value = "ARIEL ESPINOZA"
                    elif col == 6:
                        cell_new.value = 7
                    elif col in (10, 14): # MAI, ATENCIÓNABIERTA defaults
                        cell_new.value = 1
                    elif col >= 21 and col <= 200:
                        cell_new.value = 0 # Default inactive for procedures
        
        # Now fill in user values
        # Date
        fecha_str = record_data.get("fecha")
        if fecha_str:
            try:
                dt = datetime.datetime.strptime(fecha_str, "%Y-%m-%d").date()
                sheet.cell(row=target_row, column=1).value = dt
            except:
                sheet.cell(row=target_row, column=1).value = fecha_str
                
        # Ficha Usuario
        ficha_val = record_data.get("fichaUsuario")
        if ficha_val:
            try:
                sheet.cell(row=target_row, column=3).value = int(ficha_val)
            except:
                sheet.cell(row=target_row, column=3).value = ficha_val
                
        # RUT
        rut_val = record_data.get("rutUsuario")
        if rut_val:
            sheet.cell(row=target_row, column=4).value = rut_val
        elif is_new_row:
            sheet.cell(row=target_row, column=4).value = "=VLOOKUP(Tabla1[[#This Row],[FICHA USUARIO]],USUARIOS!$A:$F,2,0)"
            
        # Diagnóstico
        diag_val = record_data.get("diagnostico")
        if diag_val:
            sheet.cell(row=target_row, column=5).value = diag_val
        elif is_new_row:
            sheet.cell(row=target_row, column=5).value = "=VLOOKUP(Tabla1[[#This Row],[FICHA USUARIO]],USUARIOS!$A:$F,3,0)"
            
        # Sexo
        sexo_val = record_data.get("sexo")
        if sexo_val:
            sheet.cell(row=target_row, column=8).value = sexo_val
        elif is_new_row:
            sheet.cell(row=target_row, column=8).value = "=VLOOKUP(Tabla1[[#This Row],[FICHA USUARIO]],USUARIOS!$A:$F,4,0)"
            
        # Edad
        edad_val = record_data.get("edad")
        if edad_val:
            try:
                sheet.cell(row=target_row, column=9).value = int(edad_val)
            except:
                sheet.cell(row=target_row, column=9).value = edad_val
        elif is_new_row:
            sheet.cell(row=target_row, column=9).value = "=VLOOKUP(Tabla1[[#This Row],[FICHA USUARIO]],USUARIOS!$A:$F,5,0)"
            
        # Booleans (MAI, Pensionado, noBeneficiario, atencionCerrada, etc.)
        def set_bool_col(key, col):
            val = record_data.get(key)
            if val is not None:
                sheet.cell(row=target_row, column=col).value = 1 if val else 0

        set_bool_col("mai", 10)
        set_bool_col("pensionado", 11)
        set_bool_col("noBeneficiario", 12)
        set_bool_col("atencionCerrada", 13)
        set_bool_col("atencionAbierta", 14)
        set_bool_col("atencionUrgencia", 15)
        set_bool_col("ingreso", 16)
        set_bool_col("pti", 17)
        set_bool_col("consultaInicial", 18)
        set_bool_col("consultaIntermedia", 19)
        
        # Sessions
        sesiones_val = record_data.get("sesiones")
        if sesiones_val is not None:
            try:
                sheet.cell(row=target_row, column=20).value = int(sesiones_val)
            except:
                sheet.cell(row=target_row, column=20).value = sesiones_val
                
        # Write procedures
        for key, col in COL_MAPPING.items():
            if col > 20: # Procedure columns
                val = record_data.get(key)
                if val is not None:
                    sheet.cell(row=target_row, column=col).value = 1 if val else 0
        
        # Save workbook
        wb.save(EXCEL_PATH)
        return {"success": True, "row": target_row}
    except PermissionError:
        return {"error": "El archivo Excel está abierto por otro programa (como Excel). Por favor ciérralo e intenta de nuevo."}
    except Exception as e:
        return {"error": str(e)}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "No action specified"}))
        sys.exit(1)
        
    action = sys.argv[1]
    
    if action == "read":
        # Read last 50 records
        n = 50
        if len(sys.argv) >= 3:
            try:
                n = int(sys.argv[2])
            except:
                pass
        result = read_records(n)
        print(json.dumps(result))
        
    elif action == "search":
        if len(sys.argv) < 3:
            print(json.dumps({"error": "No search query specified"}))
            sys.exit(1)
        query = sys.argv[2]
        result = search_patient(query)
        print(json.dumps(result))
        
    elif action == "write":
        try:
            input_data = json.load(sys.stdin)
            result = write_record(input_data)
            print(json.dumps(result))
        except Exception as e:
            print(json.dumps({"error": "Failed to parse input JSON: " + str(e)}))
            sys.exit(1)
    else:
        print(json.dumps({"error": "Unknown action: " + action}))
        sys.exit(1)
